import sys
import csv
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

def main():
    csv_file = sys.argv[1] if len(sys.argv) > 1 else "results.csv"
    last_name = sys.argv[2] if len(sys.argv) > 2 else "LastName"
    out_file = f"primes_benchmark_{last_name}.xlsx"

    # ── Read CSV ──────────────────────────────
    rows = []
    with open(csv_file) as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    wb = Workbook()

    # ─────────────────────────────────────────
    # Sheet 1: Raw Results
    # ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Benchmark Results"

    # Styles
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="2F5496")
    sub_font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    mt_fill       = PatternFill("solid", start_color="4472C4")
    st_fill       = PatternFill("solid", start_color="ED7D31")
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")
    thin          = Side(style="thin", color="AAAAAA")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font     = Font(name="Arial", size=10)
    avg_font      = Font(name="Arial", size=10, bold=True)
    avg_fill_mt   = PatternFill("solid", start_color="D9E1F2")
    avg_fill_st   = PatternFill("solid", start_color="FCE4D6")

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Prime Counting Benchmark — {last_name}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["Program", "N", "Run 1 (s)", "Run 2 (s)", "Run 3 (s)", "Average (s)", "Notes"]
    ws.append([""] * 7)   # blank row 2
    ws.append(headers)    # row 3
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 20

    # Data rows
    row_num = 4
    for r in rows:
        is_mt = r["Program"] == "Multi-Threaded"
        fill  = PatternFill("solid", start_color=("EBF3FB" if is_mt else "FEF0E7"))
        afill = avg_fill_mt if is_mt else avg_fill_st

        cells = [
            r["Program"],
            int(r["N"]),
            float(r["Run1(s)"]),
            float(r["Run2(s)"]),
            float(r["Run3(s)"]),
            f"=AVERAGE(C{row_num}:E{row_num})",
            ""
        ]
        ws.append(cells)
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.font   = avg_font if col == 6 else data_font
            cell.fill   = afill   if col == 6 else fill
            cell.border = border
            cell.alignment = center if col != 1 else left
            if col in (3, 4, 5, 6):
                cell.number_format = "0.000000"
        row_num += 1

    # Column widths
    widths = {"A": 20, "B": 8, "C": 14, "D": 14, "E": 14, "F": 14, "G": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ─────────────────────────────────────────
    # Sheet 2: Comparison Summary
    # ─────────────────────────────────────────
    ws2 = wb.create_sheet("Comparison Summary")

    # Title
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "Multi-Threaded vs Single-Threaded — Average Elapsed Time"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13)
    ws2["A1"].alignment = center
    ws2.row_dimensions[1].height = 26

    # Headers
    sum_headers = ["N", "MT Avg (s)", "ST Avg (s)", "Speedup (MT/ST)", "Faster?"]
    ws2.append([""] * 5)
    ws2.append(sum_headers)
    for c, _ in enumerate(sum_headers, start=1):
        cell = ws2.cell(row=3, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ns = [1000, 2000, 4000, 8000]
    mt_rows = [r for r in rows if r["Program"] == "Multi-Threaded"]
    st_rows = [r for r in rows if r["Program"] == "Single-Threaded"]

    data_start = 4
    for i, n in enumerate(ns):
        mt_avg = float(mt_rows[i]["Average(s)"]) if i < len(mt_rows) else 0
        st_avg = float(st_rows[i]["Average(s)"]) if i < len(st_rows) else 0
        srow = data_start + i

        ws2.append([
            n,
            mt_avg,
            st_avg,
            f"=C{srow}/B{srow}" if mt_avg != 0 else "N/A",
            f'=IF(B{srow}<C{srow},"MT Faster","ST Faster")'
        ])

        row_fill = PatternFill("solid", start_color=("EBF3FB" if i % 2 == 0 else "FFFFFF"))
        for col in range(1, 6):
            cell = ws2.cell(row=srow, column=col)
            cell.font   = data_font
            cell.fill   = row_fill
            cell.border = border
            cell.alignment = center
            if col in (2, 3):
                cell.number_format = "0.000000"
            if col == 4:
                cell.number_format = "0.00x"

    for col, w in zip("ABCDE", [10, 14, 14, 18, 14]):
        ws2.column_dimensions[col].width = w

    # ── Bar Chart ─────────────────────────────
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Average Execution Time by N"
    chart.y_axis.title = "Time (seconds)"
    chart.x_axis.title = "N"
    chart.style   = 10
    chart.width   = 18
    chart.height  = 12

    mt_data = Reference(ws2, min_col=2, min_row=3, max_row=3 + len(ns))
    st_data = Reference(ws2, min_col=3, min_row=3, max_row=3 + len(ns))
    cats    = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(ns))

    chart.add_data(mt_data, titles_from_data=True)
    chart.add_data(st_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "4472C4"
    chart.series[1].graphicalProperties.solidFill = "ED7D31"

    ws2.add_chart(chart, "A9")

    wb.save(out_file)
    print(f"Excel file saved: {out_file}")

if __name__ == "__main__":
    main()
